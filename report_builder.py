"""
report_builder.py
-----------------
מייצר Excel סיכום לאחר כל ריצה — לצורך בדיקה ואימות.

גיליונות:
  1. סיכום     — שורה לכל קבוצת מייל (draft_id, נמען, מס' רשומות, טווח שבועות)
  2. פירוט     — שורה לכל רשומה (ת.ז., קוד שגיאה, שבועות, לאיזה מייל שויכה)
  3. מוחרגות   — רשומות שסוננו עם סיבה

שימוש:
  from report_builder import build_run_report
  xlsx_bytes = build_run_report(groups, send_results, skipped_records, run_date)
"""

import io
from collections import Counter, defaultdict
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def build_run_report(groups, send_results, skipped_records=None, raw_records=None, run_date=None):
    """
    groups          : מ-group_records()
    send_results    : מ-send_all_groups()  [רשימת SendResult dicts]
    skipped_records : רשימת (record, reason) שנסוננו ב-classify_all (optional)
    raw_records     : רשימת הרשומות הגולמיות מה-API (optional) — לגיליון מעקב pipeline
    run_date        : datetime או None → now()

    מחזיר bytes של Excel.
    """
    run_date = run_date or datetime.now()

    # --- draft_id lookup ---
    draft_map = {r["group_key"]: r for r in (send_results or [])}

    # === גיליון 1: סיכום ===
    summary_rows = []
    for g in groups:
        key    = g["group_key"]
        fmt    = g["email_format"]
        recs   = g["records"]
        meta   = g.get("meta", {})
        sr     = draft_map.get(key, {})

        counters = [r.get("counter") for r in recs if r.get("counter") is not None]
        c_min = min(counters) if counters else None
        c_max = max(counters) if counters else None

        summary_rows.append({
            "פורמט":           fmt,
            "מפתח קבוצה":      key,
            "גוף מוסדי / מעסיק": (
                meta.get("fund_institution_name") or
                meta.get("customer_name") or
                meta.get("customer_number") or ""
            ),
            "ח.פ / מזהה":      meta.get("customer_number") or meta.get("fund_institution_id") or "",
            "מס' רשומות":      len(recs),
            "שבועות (מינ)":    c_min,
            "שבועות (מקס)":    c_max,
            "draft_id":        sr.get("draft_id"),
            "נשלח בהצלחה":     "כן" if sr.get("ok") else ("לא" if sr else "—"),
            "שגיאה":           sr.get("error") or "",
        })

    # === גיליון 2: פירוט ===
    detail_rows = []
    for g in groups:
        key = g["group_key"]
        fmt = g["email_format"]
        sr  = draft_map.get(key, {})
        for r in g["records"]:
            detail_rows.append({
                "פורמט":           fmt,
                "מפתח קבוצה":      key,
                "draft_id":        sr.get("draft_id"),
                "record_id":       r.get("record_id"),
                "ח.פ מעסיק":       r.get("customer_number"),
                "שם מעסיק":        r.get("customer_name"),
                "מ.ז. עובד":       r.get("employee_id"),
                "שם עובד":         r.get("full_name"),
                "קוד שגיאה":       r.get("error_code"),
                "תיאור שגיאה":     r.get("error_description"),
                "שבועות":          r.get("counter"),
                "אחריות":          r.get("responsibility"),
                "נתיב ניתוב":      r.get("routing_path"),
                "גוף מוסדי":       r.get("fund_institution_name"),
                "שם קובץ מקור":    r.get("original_file_name"),
                "חודש שכר":        r.get("_raw", {}).get("CHODESH_MASKORET"),
            })

    # === גיליון 3: מוחרגות ===
    skipped_rows = []
    for item in (skipped_records or []):
        if isinstance(item, tuple) and len(item) == 2:
            rec, reason = item
        else:
            rec, reason = item, "סונן"
        skipped_rows.append({
            "record_id":   rec.get("MISPAR_MEZAHE_RESHUMA") or rec.get("record_id"),
            "ח.פ מעסיק":  rec.get("CustomerNumber") or rec.get("customer_number"),
            "קוד שגיאה":  rec.get("ErrorCodeV4Id") or rec.get("error_code"),
            "שבועות":     rec.get("OnlyOnStatusChange_DatesDiffInWeeks") or rec.get("counter"),
            "סיבה":       reason,
        })

    # === בניית Excel ===
    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        pd.DataFrame(summary_rows).to_excel(writer, sheet_name="סיכום",   index=False)
        pd.DataFrame(detail_rows).to_excel(writer,  sheet_name="פירוט",   index=False)
        pd.DataFrame(skipped_rows).to_excel(writer, sheet_name="מוחרגות", index=False)

    bio.seek(0)
    wb = load_workbook(bio)
    _style_workbook(wb, run_date)
    _build_dashboard_sheet(wb, groups, skipped_records or [], run_date)
    if raw_records:
        _build_pipeline_sheet(wb, raw_records, groups, skipped_records or [], draft_map)

    # הזזת דשבורד להיות הגיליון הראשון
    wb.move_sheet("דשבורד", offset=-wb.sheetnames.index("דשבורד"))

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _style_workbook(wb, run_date):
    header_fill = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    data_font   = Font(name="Arial", size=10)
    center      = Alignment(horizontal="center", vertical="center", wrap_text=False)
    right       = Alignment(horizontal="right",  vertical="center", wrap_text=False)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # כותרת גיליון
        ws.sheet_view.rightToLeft = True

        # עיצוב header
        for cell in ws[1]:
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = center

        # עיצוב data
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font      = data_font
                cell.alignment = right

        # רוחב עמודות אוטומטי
        for col_cells in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 4, 50)

        ws.freeze_panes = "A2"


def _build_dashboard_sheet(wb, groups, skipped_records, run_date):
    """מוסיף גיליון דשבורד עם 4 טבלאות סיכום."""
    ws = wb.create_sheet("דשבורד")
    ws.sheet_view.rightToLeft = True

    H_FILL  = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
    H_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    T_FILL  = PatternFill("solid", start_color="D6E4F0", end_color="D6E4F0")  # כותרת טבלה
    T_FONT  = Font(bold=True, name="Arial", size=10)
    D_FONT  = Font(name="Arial", size=10)
    RIGHT   = Alignment(horizontal="right", vertical="center")
    CENTER  = Alignment(horizontal="center", vertical="center")

    def _hcell(ws, row, col, value):
        c = ws.cell(row=row, column=col, value=value)
        c.font = H_FONT; c.fill = H_FILL; c.alignment = CENTER

    def _tcell(ws, row, col, value):
        c = ws.cell(row=row, column=col, value=value)
        c.font = T_FONT; c.fill = T_FILL; c.alignment = RIGHT

    def _dcell(ws, row, col, value):
        c = ws.cell(row=row, column=col, value=value)
        c.font = D_FONT; c.alignment = RIGHT

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 14

    # ---- נתונים גולמיים ----
    all_records = [r for g in groups for r in g["records"]]
    total_treated = len(all_records)
    skipped_list  = list(skipped_records)
    total_skipped = len(skipped_list)
    total_fetched = total_treated + total_skipped

    # ספירת סיבות סינון
    skip_counts = Counter()
    for _, reason in skipped_list:
        r = reason or "אחר"
        if "Counter=0" in r:
            skip_counts["Counter=0 (שגיאה חדשה)"] += 1
        elif "קוד שגיאה 1" in r or "קוד שגיאה 2" in r:
            skip_counts["קוד שגיאה 1/2 (מוחרג)"] += 1
        elif "מבוטלת" in r:
            skip_counts["רשומה מבוטלת"] += 1
        elif "מוחרג בקובץ מיפוי" in r:
            skip_counts["מוחרג בקובץ מיפוי"] += 1
        else:
            skip_counts["אחר"] += 1

    # ---- טבלה 1: סיכום כללי (שורה 1) ----
    row = 1
    _hcell(ws, row, 1, f"דשבורד — ריצה {run_date.strftime('%d/%m/%Y %H:%M')} UTC")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 2

    _tcell(ws, row, 1, "קטגוריה"); _tcell(ws, row, 2, "כמות"); _tcell(ws, row, 3, "אחוז")
    row += 1
    for label, val in [
        ("סה\"כ רשומות שהתקבלו",  total_fetched),
        ("רשומות לטיפול",         total_treated),
        ("רשומות לא לטיפול",      total_skipped),
        ("  — Counter=0 (שגיאה חדשה)",      skip_counts.get("Counter=0 (שגיאה חדשה)", 0)),
        ("  — קוד שגיאה 1/2 (מוחרג)",       skip_counts.get("קוד שגיאה 1/2 (מוחרג)", 0)),
        ("  — רשומה מבוטלת",                 skip_counts.get("רשומה מבוטלת", 0)),
        ("  — מוחרג בקובץ מיפוי",            skip_counts.get("מוחרג בקובץ מיפוי", 0)),
        ("  — אחר",                           skip_counts.get("אחר", 0)),
    ]:
        pct = f"{val/total_fetched*100:.1f}%" if total_fetched else "—"
        _dcell(ws, row, 1, label); _dcell(ws, row, 2, val); _dcell(ws, row, 3, pct)
        row += 1
    row += 1

    # ---- טבלה 2: לפי גורם אחראי (שורה row) ----
    _tcell(ws, row, 1, "גורם אחראי"); _tcell(ws, row, 2, "כמות רשומות"); _tcell(ws, row, 3, "כמות מיילים")
    row += 1
    by_format = defaultdict(list)
    for g in groups:
        by_format[g["email_format"]].append(g)
    for fmt, grps in sorted(by_format.items()):
        rec_count  = sum(len(g["records"]) for g in grps)
        mail_count = len(grps)
        _dcell(ws, row, 1, fmt); _dcell(ws, row, 2, rec_count); _dcell(ws, row, 3, mail_count)
        row += 1
    row += 1

    # ---- טבלה 3: לפי מדיניות הסלמה (counter) ----
    _tcell(ws, row, 1, "שבועות (Counter)"); _tcell(ws, row, 2, "כמות רשומות"); _tcell(ws, row, 3, "אחוז מטופלים")
    row += 1
    counter_counts = Counter()
    for r in all_records:
        c = r.get("counter")
        try:
            c = int(float(c)) if c is not None else 0
        except (ValueError, TypeError):
            c = 0
        bucket = str(c) if c <= 4 else "5+"
        counter_counts[bucket] += 1
    for bucket in ["1", "2", "3", "4", "5+"]:
        val = counter_counts.get(bucket, 0)
        pct = f"{val/total_treated*100:.1f}%" if total_treated else "—"
        _dcell(ws, row, 1, f"שבוע {bucket}"); _dcell(ws, row, 2, val); _dcell(ws, row, 3, pct)
        row += 1
    row += 1

    # ---- טבלה 4: טופ 20 מעסיקים לפי הסלמה ----
    headers = ["מעסיק", "ח.פ", "סה\"כ", "שבוע 1", "שבוע 2", "שבוע 3", "שבוע 4", "שבוע 5+"]
    for i, h in enumerate(headers, 1):
        _tcell(ws, row, i, h)
    row += 1

    employer_data = defaultdict(lambda: Counter())
    employer_names = {}
    for r in all_records:
        cnum = str(r.get("customer_number") or "—")
        employer_names[cnum] = r.get("customer_name") or r.get("employer_name") or cnum
        c = r.get("counter")
        try:
            c = int(float(c)) if c is not None else 0
        except (ValueError, TypeError):
            c = 0
        bucket = str(c) if c <= 4 else "5+"
        employer_data[cnum][bucket] += 1

    top20 = sorted(employer_data.items(), key=lambda x: sum(x[1].values()), reverse=True)[:20]
    for cnum, buckets in top20:
        total = sum(buckets.values())
        row_vals = [
            employer_names.get(cnum, cnum), cnum, total,
            buckets.get("1", 0), buckets.get("2", 0),
            buckets.get("3", 0), buckets.get("4", 0), buckets.get("5+", 0),
        ]
        for i, v in enumerate(row_vals, 1):
            _dcell(ws, row, i, v)
        row += 1


def _build_pipeline_sheet(wb, raw_records, groups, skipped_records, draft_map):
    """גיליון מעקב pipeline — שורה לכל רשומה גולמית מה-API עם גורל הרשומה."""

    # --- בניית lookups ---
    # record_id → (classified_record, group_key, draft_id)
    classified_lookup = {}
    for g in groups:
        gk  = g["group_key"]
        fmt = g["email_format"]
        sr  = draft_map.get(gk, {})
        did = sr.get("draft_id")
        for r in g["records"]:
            classified_lookup[r["record_id"]] = {
                "email_format":  fmt,
                "responsibility": r.get("responsibility"),
                "group_key":     gk,
                "draft_id":      did,
                "routing_path":  r.get("routing_path"),
            }

    # record_id → reason (סונן)
    skipped_lookup = {}
    for raw_rec, reason in skipped_records:
        rid = raw_rec.get("MISPAR_MEZAHE_RESHUMA") or raw_rec.get("record_id")
        if rid:
            skipped_lookup[str(rid)] = reason or "סונן"

    # --- בניית שורות ---
    rows = []
    for raw in raw_records:
        rid = str(raw.get("MISPAR_MEZAHE_RESHUMA") or "")
        cl  = classified_lookup.get(rid)
        sk  = skipped_lookup.get(rid)

        if cl:
            action = "טופל"
            reason = ""
            fmt    = cl["email_format"]
            resp   = cl["responsibility"]
            gk     = cl["group_key"]
            did    = cl["draft_id"]
            path   = cl["routing_path"]
        elif sk is not None:
            action = "סונן"
            reason = sk
            fmt = resp = gk = did = path = ""
        else:
            action = "לא ידוע"
            reason = fmt = resp = gk = did = path = ""

        rows.append({
            "MISPAR_MEZAHE_RESHUMA":              rid,
            "CustomerNumber":                     raw.get("CustomerNumber"),
            "CustomerName":                       raw.get("CustomerName"),
            "ErrorCodeV4Id":                      raw.get("ErrorCodeV4Id"),
            "StatusDescription":                  raw.get("StatusDescription"),
            "OnlyOnStatusChange_DatesDiffInWeeks": raw.get("OnlyOnStatusChange_DatesDiffInWeeks"),
            "LastPositive_CHODESH_MASKORET":      raw.get("LastPositive_CHODESH_MASKORET"),
            "FundInstitutionName":                raw.get("FundInstitutionName"),
            "CHODESH_MASKORET":                   raw.get("CHODESH_MASKORET"),
            "פעולה":         action,
            "סיבת סינון":    reason,
            "email_format":  fmt,
            "אחריות":        resp,
            "group_key":     gk,
            "draft_id":      did,
            "routing_path":  path,
        })

    ws_name = "מעקב pipeline"
    df = pd.DataFrame(rows)

    # כתיבה זמנית דרך BytesIO כדי להוסיף לחוברת הקיימת
    tmp = io.BytesIO()
    with pd.ExcelWriter(tmp, engine="openpyxl") as w:
        df.to_excel(w, sheet_name=ws_name, index=False)
    tmp.seek(0)
    src_wb = load_workbook(tmp)
    src_ws = src_wb[ws_name]

    # העברת הגיליון לחוברת הראשית
    tgt_ws = wb.create_sheet(ws_name)
    tgt_ws.sheet_view.rightToLeft = True
    for row in src_ws.iter_rows(values_only=True):
        tgt_ws.append(list(row))

    # עיצוב בסיסי
    H_FILL = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
    H_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    D_FONT = Font(name="Arial", size=10)
    RIGHT  = Alignment(horizontal="right", vertical="center")
    CENTER = Alignment(horizontal="center", vertical="center")

    for cell in tgt_ws[1]:
        cell.font = H_FONT; cell.fill = H_FILL; cell.alignment = CENTER
    for row in tgt_ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = D_FONT; cell.alignment = RIGHT

    for col_cells in tgt_ws.columns:
        max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
        tgt_ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 4, 40)

    tgt_ws.freeze_panes = "A2"
